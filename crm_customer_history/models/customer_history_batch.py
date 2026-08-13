import base64
import csv
import io
import re
import zipfile
import unicodedata
from datetime import date, datetime

import openpyxl
import xlsxwriter
from openpyxl.utils.exceptions import InvalidFileException

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError


class CrmCustomerHistoryBatch(models.Model):
    _name = 'crm.customer.history.batch'
    _description = 'Lote de históricos comerciales'
    _rec_name = 'name'
    _order = 'create_date desc, id desc'
    _MAX_FILE_SIZE = 10 * 1024 * 1024

    name = fields.Char(string='Nombre del lote', required=True, default='Nuevo histórico')
    company_id = fields.Many2one(
        'res.company', string='Empresa', required=True, default=lambda self: self.env.company,
        index=True)
    file_data = fields.Binary(string='Archivo CSV/XLSX', attachment=True, copy=False)
    file_name = fields.Char(string='Nombre del archivo', copy=False)
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('loaded', 'Cargado'),
        ('approved', 'Aprobado'),
        ('cancelled', 'Cancelado'),
    ], string='Estado', default='draft', required=True, index=True)
    create_missing_partners = fields.Boolean(
        string='Crear clientes no encontrados', default=True,
        help='Crea un contacto básico cuando el histórico no coincide por RUC, correo, teléfono o nombre.')
    line_ids = fields.One2many('crm.customer.history.line', 'batch_id', string='Detalle')
    line_count = fields.Integer(string='Filas', compute='_compute_counts')
    valid_count = fields.Integer(string='Válidas', compute='_compute_counts')
    error_count = fields.Integer(string='Errores', compute='_compute_counts')
    duplicate_count = fields.Integer(string='Duplicados', compute='_compute_counts')
    unmatched_count = fields.Integer(string='Sin identificar', compute='_compute_counts')
    history_date_from = fields.Date(string='Desde', compute='_compute_date_range', store=True)
    history_date_to = fields.Date(string='Hasta', compute='_compute_date_range', store=True)
    note = fields.Text(string='Notas')
    approved_at = fields.Datetime(string='Aprobado el', readonly=True, copy=False)
    approved_by = fields.Many2one('res.users', string='Aprobado por', readonly=True, copy=False)
    cancelled_at = fields.Datetime(string='Cancelado el', readonly=True, copy=False)
    cancelled_by = fields.Many2one('res.users', string='Cancelado por', readonly=True, copy=False)

    @api.depends('line_ids.state', 'line_ids.partner_match_state', 'line_ids.invoice_date')
    def _compute_counts(self):
        for batch in self:
            lines = batch.line_ids
            batch.line_count = len(lines)
            batch.valid_count = len(lines.filtered(lambda line: line.state == 'valid'))
            batch.error_count = len(lines.filtered(lambda line: line.state == 'error'))
            batch.duplicate_count = len(lines.filtered(lambda line: line.state == 'duplicate'))
            batch.unmatched_count = len(lines.filtered(lambda line: line.partner_match_state == 'unmatched'))

    @api.depends('line_ids.invoice_date')
    def _compute_date_range(self):
        for batch in self:
            dates = batch.line_ids.mapped('invoice_date')
            batch.history_date_from = min(dates) if dates else False
            batch.history_date_to = max(dates) if dates else False

    @api.constrains('file_name')
    def _check_file_name(self):
        for batch in self:
            if batch.file_name and not batch.file_name.lower().endswith(('.csv', '.xlsx')):
                raise ValidationError(_('El archivo debe ser CSV o XLSX.'))

    @api.constrains('file_data')
    def _check_file_size(self):
        for batch in self:
            if batch.file_data and len(base64.b64decode(batch.file_data)) > self._MAX_FILE_SIZE:
                raise ValidationError(_('El archivo no puede superar los 10 MB.'))

    @staticmethod
    def _normalize_header(value):
        text = unicodedata.normalize('NFKD', str(value or ''))
        text = ''.join(char for char in text if not unicodedata.combining(char))
        return re.sub(r'[^a-z0-9]+', '_', text.lower()).strip('_')

    @staticmethod
    def _text(value):
        return str(value or '').strip()

    @staticmethod
    def _parse_amount(value):
        if value in (None, ''):
            raise ValueError(_('Falta el monto.'))
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace('$', '').replace('€', '').replace(' ', '')
        if ',' in text and '.' in text:
            if text.rfind(',') > text.rfind('.'):
                text = text.replace('.', '').replace(',', '.')
            else:
                text = text.replace(',', '')
        elif ',' in text:
            text = text.replace(',', '.')
        amount = float(text)
        if amount <= 0:
            raise ValueError(_('El monto debe ser mayor que cero.'))
        return amount

    @staticmethod
    def _parse_date(value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value or '').strip()
        if not text:
            raise ValueError(_('Falta la fecha.'))
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d'):
            try:
                return datetime.strptime(text[:10], fmt).date()
            except ValueError:
                continue
        raise ValueError(_('Fecha no reconocida: %s') % text)

    def _read_rows(self):
        self.ensure_one()
        raw = base64.b64decode(self.file_data or b'')
        if (self.file_name or '').lower().endswith('.xlsx'):
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            except (InvalidFileException, zipfile.BadZipFile, OSError, ValueError) as error:
                raise UserError(_('No se pudo leer el archivo XLSX: %s') % error) from error
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                return []
            return [dict(zip(headers, row)) for row in rows if any(value not in (None, '') for value in row)]
        text = raw.decode('utf-8-sig', errors='replace')
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=',;\t')
        except csv.Error:
            dialect = csv.excel
        return list(csv.DictReader(io.StringIO(text), dialect=dialect))

    @api.model
    def _aliases(self):
        return {
            'customer_name': {'nombre_cliente', 'cliente', 'nombre', 'customer_name', 'customer', 'name'},
            'customer_external_ref': {'codigo_cliente', 'codigo_externo', 'referencia', 'ref', 'customer_code'},
            'customer_vat': {'ruc', 'cedula', 'identificacion', 'identificacion_cliente', 'vat', 'tax_id'},
            'customer_email': {'correo', 'email', 'correo_electronico', 'customer_email'},
            'customer_phone': {'telefono', 'celular', 'phone', 'mobile'},
            'invoice_number': {'factura', 'numero_factura', 'n_factura', 'invoice', 'invoice_number', 'documento'},
            'invoice_date': {'fecha', 'fecha_factura', 'date', 'invoice_date', 'fecha_compra'},
            'amount_total': {'monto', 'monto_total', 'total', 'importe', 'amount', 'amount_total', 'valor'},
            'currency': {'moneda', 'currency', 'divisa'},
            'move_type': {'tipo', 'tipo_movimiento', 'movement_type', 'move_type'},
        }

    def _map_row(self, row):
        normalized = {self._normalize_header(key): value for key, value in row.items() if key is not None}
        result = {}
        for target, aliases in self._aliases().items():
            for alias in aliases:
                if alias in normalized and normalized[alias] not in (None, ''):
                    result[target] = normalized[alias]
                    break
        return result

    def _find_partner(self, values):
        Partner = self.env['res.partner'].with_company(self.company_id)
        vat = self._text(values.get('customer_vat'))
        email = self._text(values.get('customer_email')).lower()
        phone = self.env['crm.customer.history.line'].normalize_phone(values.get('customer_phone'))
        name = self._text(values.get('customer_name'))
        if vat:
            partner = Partner.search([('vat', '=ilike', vat)], limit=1)
            if partner:
                return partner, 'matched'
        if email:
            partner = Partner.search([('email', '=ilike', email)], limit=1)
            if partner:
                return partner, 'matched'
        if phone:
            candidates = Partner.search(['|', ('phone', '!=', False), ('mobile', '!=', False)])
            for partner in candidates:
                if phone in self.env['crm.customer.history.line'].normalize_phone(partner.phone) or phone in self.env['crm.customer.history.line'].normalize_phone(partner.mobile):
                    return partner, 'matched'
        if name:
            partner = Partner.search([('name', '=ilike', name)], limit=1)
            if partner:
                return partner, 'matched'
        if self.create_missing_partners and name:
            vals = {'name': name}
            if vat:
                vals['vat'] = vat
            if email:
                vals['email'] = email
            if values.get('customer_phone'):
                vals['phone'] = self._text(values['customer_phone'])
            return Partner.create(vals), 'created'
        return self.env['res.partner'], 'unmatched'

    def _resolve_currency(self, value):
        text = self._text(value).upper()
        if not text:
            return self.company_id.currency_id
        return self.env['res.currency'].search([
            '|', ('name', '=ilike', text), ('symbol', '=ilike', text),
        ], limit=1) or self.company_id.currency_id

    def _build_line_values(self, values, row_number):
        mapped = self._map_row(values)
        invoice_date = self._parse_date(mapped.get('invoice_date'))
        if invoice_date > fields.Date.context_today(self):
            raise ValueError(_('La fecha de factura no puede estar en el futuro.'))
        amount = self._parse_amount(mapped.get('amount_total'))
        move_type = 'refund' if self._normalize_header(mapped.get('move_type')) in {'refund', 'devolucion', 'devoluciones'} else 'sale'
        partner, match_state = self._find_partner(mapped)
        line_values = {
            'batch_id': self.id,
            'company_id': self.company_id.id,
            'row_number': row_number,
            'customer_name': self._text(mapped.get('customer_name')),
            'customer_external_ref': self._text(mapped.get('customer_external_ref')),
            'customer_vat': self._text(mapped.get('customer_vat')),
            'customer_email': self._text(mapped.get('customer_email')),
            'customer_phone': self._text(mapped.get('customer_phone')),
            'partner_id': partner.id if partner else False,
            'partner_match_state': match_state,
            'invoice_number': self._text(mapped.get('invoice_number')) or _('Sin número'),
            'invoice_date': invoice_date,
            'move_type': move_type,
            'amount_total': amount,
            'currency_id': self._resolve_currency(mapped.get('currency')).id,
            'state': 'valid' if partner else 'error',
            'error_message': False if partner else _('No se encontró el cliente y está desactivada la creación automática.'),
        }
        line_values['fingerprint'] = self.env['crm.customer.history.line'].make_fingerprint(line_values)
        duplicate = self.env['crm.customer.history.line'].search([('fingerprint', '=', line_values['fingerprint'])], limit=1)
        if duplicate:
            line_values['state'] = 'duplicate'
            line_values['error_message'] = _('Duplicado del lote %s, fila %s.') % (duplicate.batch_id.display_name, duplicate.row_number)
        return line_values

    def action_import_file(self):
        self.ensure_one()
        if self.state == 'approved':
            raise UserError(_('Un lote aprobado no se puede volver a importar. Cree un lote nuevo.'))
        if not self.file_data or not self.file_name:
            raise UserError(_('Adjunte un archivo CSV o XLSX antes de importar.'))
        self.line_ids.sudo().unlink()
        for row_number, row in enumerate(self._read_rows(), 2):
            try:
                values = self._build_line_values(row, row_number)
            except (ValueError, TypeError, UserError) as error:
                values = {
                    'batch_id': self.id,
                    'company_id': self.company_id.id,
                    'row_number': row_number,
                    'customer_name': self._text(row.get('Nombre Cliente') or row.get('Cliente') or row.get('nombre')),
                    'state': 'error',
                    'error_message': str(error),
                }
            self.env['crm.customer.history.line'].create(values)
        self.state = 'loaded'
        if not self.line_ids:
            raise UserError(_('El archivo no contiene filas para importar.'))
        return self.action_open_lines()

    def action_revalidate(self):
        for batch in self:
            for line in batch.line_ids.filtered(lambda item: item.state == 'error' and item.partner_id):
                line.write({'state': 'valid', 'error_message': False, 'partner_match_state': 'matched'})
            batch.state = 'loaded'
        return True

    def action_approve(self):
        for batch in self:
            if not batch.valid_count:
                raise UserError(_('No hay filas válidas para aprobar. Corrija los errores primero.'))
            batch.write({
                'state': 'approved',
                'approved_at': fields.Datetime.now(),
                'approved_by': self.env.user.id,
                'cancelled_at': False,
                'cancelled_by': False,
            })
        self.env['res.partner'].sudo()._cron_compute_rfm_scores()
        return True

    def action_cancel(self):
        if any(batch.state == 'approved' for batch in self):
            raise UserError(_(
                'Un lote aprobado debe revertirse con la acción "Revertir aprobación".'))
        self.write({
            'state': 'cancelled',
            'cancelled_at': fields.Datetime.now(),
            'cancelled_by': self.env.user.id,
        })
        return True

    def action_revert_approval(self):
        """Retira un lote aprobado sin borrar sus datos ni las evidencias."""
        if any(batch.state != 'approved' for batch in self):
            raise UserError(_('Solo se pueden revertir lotes aprobados.'))
        self.write({
            'state': 'cancelled',
            'cancelled_at': fields.Datetime.now(),
            'cancelled_by': self.env.user.id,
        })
        self.env['res.partner'].sudo()._cron_compute_rfm_scores()
        return True

    def action_open_lines(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Detalle del histórico'),
            'res_model': 'crm.customer.history.line',
            'view_mode': 'list,form',
            'domain': [('batch_id', '=', self.id)],
            'context': {'default_batch_id': self.id},
            'target': 'current',
        }

    def action_open_errors(self):
        self.ensure_one()
        action = self.action_open_lines()
        action['domain'] = [('batch_id', '=', self.id), ('state', 'in', ('error', 'duplicate'))]
        action['name'] = _('Errores y duplicados')
        return action

    def action_export_csv(self):
        """Download the complete reviewed result, including rejected rows."""
        self.ensure_one()
        output = io.StringIO(newline='')
        writer = csv.writer(output)
        writer.writerow([
            'Fila', 'Cliente', 'RUC', 'Correo', 'Teléfono', 'Factura',
            'Fecha', 'Monto', 'Moneda', 'Tipo', 'Estado', 'Identificación',
            'Error',
        ])
        for line in self.line_ids.sorted(key=lambda item: (item.row_number, item.id)):
            writer.writerow([
                line.row_number,
                line.customer_name or '',
                line.customer_vat or '',
                line.customer_email or '',
                line.customer_phone or '',
                line.invoice_number or '',
                line.invoice_date or '',
                line.amount_total or 0.0,
                line.currency_id.name or '',
                line.move_type or '',
                line.state or '',
                line.partner_match_state or '',
                line.error_message or '',
            ])
        attachment = self.env['ir.attachment'].create({
            'name': 'resultado_%s.csv' % (self.file_name or self.name),
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue().encode('utf-8-sig')),
            'mimetype': 'text/csv',
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def action_recompute_rfm(self):
        self.env['res.partner'].sudo()._cron_compute_rfm_scores()
        return True

    def action_download_template(self):
        self.ensure_one()
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet('Historico')
        header = workbook.add_format({'bold': True, 'bg_color': '#714B67', 'font_color': 'white'})
        headers = ['Nombre Cliente', 'RUC', 'Correo', 'Teléfono', 'Número Factura', 'Fecha Factura', 'Monto', 'Moneda', 'Tipo']
        sheet.write_row(0, 0, headers, header)
        sheet.write_row(1, 0, ['Cliente de ejemplo', '9999999999', 'cliente@empresa.com', '0999999999', 'FAC-001', '2025-01-31', 100.00, self.company_id.currency_id.name, 'Venta'])
        sheet.set_column('A:A', 26); sheet.set_column('B:D', 20); sheet.set_column('E:E', 18); sheet.set_column('F:I', 16)
        workbook.close()
        attachment = self.env['ir.attachment'].create({
            'name': 'plantilla_historicos_comerciales.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': '/web/content/%s?download=true' % attachment.id, 'target': 'self'}

    @api.model
    def get_summary(self):
        batches = self.search([('state', '=', 'approved')])
        return {
            'batches': len(batches),
            'lines': sum(batches.mapped('valid_count')),
            'amount': sum(batches.mapped('line_ids.company_amount_total')),
            'customers': len(batches.mapped('line_ids.partner_id')),
        }
